import time
import json
import csv
import tempfile
import threading
import pyautogui
import pyperclip
import os
from pathlib import Path
from collections import defaultdict
from contextlib import contextmanager
from openpyxl import Workbook, load_workbook
from pynput import keyboard
from pynput.keyboard import Controller, Key
import subprocess
import sys
from tkinter import Tk, filedialog

class db:
    @staticmethod
    def d(msg):
        print(f"[DEBUG] {msg}")


# Hardcoded monitor-calibration positions, captured once via calibrate() for a
# fixed monitor/Lightroom-panel layout. Edit these directly if your monitor
# setup or panel layout changes, instead of re-running calibration every time.
# If ANY of these is None, __init__ leaves the matching self.*_pos as None,
# and run() falls back to running calibrate() (which re-prompts for and
# re-captures all five together) -- see run().
CAMERA_MAKE_POS = (2374, 488)
FILM_FORMAT_POS = (2388, 590)
SCAN_METHOD_POS = (2391, 836)
PUSH_PULL_POS = (2392, 977)
DEVELOPED_AT_POS = (2390, 997)

# Camera IDs (cameralist.xlsx's 'id' column) that don't have interchangeable
# lenses -- CAM+LNS combining in Intellectual Genre doesn't make sense for
# these, since there's only ever one lens. Add more IDs here as needed.
FIXED_LENS_CAMERA_IDS = {"RF3.5B", "ZOOM28", "UGO", "ACCZOOM", "PERKEO-II"}

# Dropdown fields open with whatever value they currently hold highlighted, not
# always the first entry -- and that value can be blank/empty, which previously
# broke selection outright (a down-count computed for "starts at entry 1" lands
# on the wrong entry, or nowhere, if the field actually starts empty). This is a
# generous upper bound on any of these controlled-vocabulary dropdowns' entry
# counts, pressed as "up" before every dropdown selection to reliably clamp to
# the first entry regardless of starting value. Edit here if a future dropdown
# ever has more entries than this.
DROPDOWN_RESET_PRESSES = 6

# Written by Importer.exportSelection() -- fixed path, same convention as
# metadataTool's vc_manifest_path, since the Lua side has no way to know
# which roll is currently being processed. Module-level (not tied to a
# metadataTool instance) so importMetadata.py can read Lightroom's current
# selection to auto-detect the roll BEFORE any xlsx/roll_root is known --
# see read_lr_selection() below.
SELECTION_MANIFEST_PATH = Path(__file__).parent / "selection_manifest.txt"


def click_plugin_extras_menu_item(item_title):
    """Clicks Library > Plug-in Extras > {item_title} via AppleScript/System Events.

    item_title must match the RENDERED menu item name, not Info.lua's source
    title -- Lightroom indents every plugin's menu items under an auto-inserted
    plugin-name header at runtime, so eg. "JSON Import" is only reachable as
    "   JSON Import" (3 leading spaces). Confirmed via a live accessibility
    dump of the actual submenu -- see git history for the full diagnosis. A
    newly added menu item's rendered indentation isn't guaranteed identical;
    re-check with a live dump if a new item's clicks silently no-op."""

    script = f'''
    tell application "Adobe Lightroom Classic" to activate
    delay 0.3
    tell application "System Events"
        tell process "Adobe Lightroom Classic"
            click menu bar item "Library" of menu bar 1
            delay 0.3
            click menu item "Plug-in Extras" of menu 1 of menu bar item "Library" of menu bar 1
            delay 0.3
            tell menu item "{item_title}" of menu 1 of menu item "Plug-in Extras" of menu 1 of menu bar item "Library" of menu bar 1
                perform action "AXPress"
            end tell
            delay 0.3
        end tell
    end tell
    '''

    return subprocess.run(["osascript", "-e", script], capture_output=True, text=True)


def read_lr_selection(timeout=5.0):
    """Triggers Lightroom's read-only "Export Selection" Plug-in Extra
    (Importer.exportSelection()) and reads back Lightroom's CURRENT selection
    -- whatever's selected at the moment this runs, no selection change is
    made here -- as an ordered list of (fileName, path) tuples.

    Module-level, not a method: used both by metadataTool.verify_selection_
    filenames() (checking the final selection against a loaded xlsx) and by
    importMetadata.py (auto-detecting which roll is currently selected,
    before any xlsx is even chosen).

    Returns None, with an explanatory message already printed, if the menu
    click fails or the manifest doesn't show up within timeout seconds."""

    SELECTION_MANIFEST_PATH.write_text("")

    result = click_plugin_extras_menu_item("   Export Selection")

    if result.returncode != 0:
        print("=" * 70)
        print("ERROR: 'Export Selection' menu click failed -- could not read back")
        print("Lightroom's current selection.")
        print(result.stderr.strip())
        print('Check that Library > Plug-in Extras > "Export Selection" exists in')
        print('the menu (Plug-in Manager > NLP Importer must be enabled).')
        print("=" * 70)
        return None

    # Read-only filename export -- no catalog-wide scan like JSON Import
    # does, so (unlike apply_lrplugin's wait-for-Enter) a short poll for
    # the file to be written is reliable instead of needing a keypress.
    deadline = time.time() + timeout
    while time.time() < deadline:
        text = SELECTION_MANIFEST_PATH.read_text().strip()
        if text:
            rows = []
            for line in text.split("\n"):
                fname, _, path = line.partition("\t")
                rows.append((fname, path))
            return rows
        time.sleep(0.1)

    print("ERROR: Timed out waiting for selection_manifest.txt from Lightroom.")
    return None


class metadataTool:

    def __init__(self, xlsx_path=None, raw_folder=None):

        self.ignore_esc = False

        self.delay_default = 0.002       # generic gap after a hotkey() combo
        self.delay_keypress = 0.002      # generic gap after a single press()
        self.delay_paste = 0.005          # gap between copy/select-all/paste within one field; unstable < ~0.01
        self.delay_clipboard_release = 0.15  # gap after a field's paste, before the next field's clipboard write; unstable < ~0.075
        self.delay_finish_image = 0.1    # settle after advancing to the next image in the per-image loop (real decode/render)
        self.delay_field_close = 0.0375  # settle after exiting field-edit mode with no image navigation (apply_shared_nlp_metadata only)
        self.delay_dropdown_deselect_settle = 0.2  # settle after apply_dropdown_fields()'s closing "up" (collapses roll-wide selection to one image)
        self.delay_dropdown_confirm = 0.4  # settle before/after confirming a dropdown entry with enter; not yet frame-trace-tuned
        self.delay_start = 0.1           # pause before the per-image loop begins
        self.delay_mousemove = 0.05      # settle after moving the mouse, before clicking

        # Runtime diagnostics: label -> [call_count, total_seconds]. Filled in by
        # _sleep()/_timed() as the run progresses, printed by print_diagnostics()
        # at the end of run() so the delays above can be tuned against real
        # measurements instead of guesswork.
        self.timings = defaultdict(lambda: [0, 0.0])
        self.images_processed = 0

        self.accept_event = threading.Event()
        # Separate from accept_event -- apply_lrplugin() waits for an actual
        # Enter press (which also dismisses Lightroom's completion dialog,
        # since Enter activates its default OK button), not the "." key used
        # for calibration capture elsewhere.
        self.enter_event = threading.Event()
        self.stop_flag = False
        self.acceptButton = "."
        self.pause_field = False
        self.pause_nextImage = False

        self.cameraMake_pos = CAMERA_MAKE_POS

        # Lightroom combo-box/dropdown fields -- unlike the free-text fields in
        # self.fields, these can't be reached by tabbing and can't be filled via
        # paste_text(); they're selected via calibrated click + up/down-arrow +
        # enter (see calibrate() / select_dropdown() / apply_dropdown_fields()).
        self.filmFormat_pos = FILM_FORMAT_POS
        self.scanMethod_pos = SCAN_METHOD_POS
        self.pushPull_pos = PUSH_PULL_POS
        self.developedAt_pos = DEVELOPED_AT_POS

        self.kb = Controller()

        pyautogui.FAILSAFE = True

        self.listener = keyboard.Listener(on_press=self._on_press)
        self.listener.start()

        self.script_dir = Path(__file__).parent

        # Allow an external caller (eg. importMetadata.py) to point this at a specific
        # roll's metadata xlsx / raw folder instead of always using the fixed
        # lrplugin-dev/metadata.xlsx. Calling metadataTool() with no args keeps the
        # original standalone behaviour unchanged.
        self.xlsx_path = Path(xlsx_path) if xlsx_path else self.script_dir / "metadata.xlsx"
        # NOTE: the Lightroom-side plugin (lrplugin-dev/data/nlp-importer.lrplugin/Importer.lua)
        # reads from a fixed, hardcoded path -- it has no way to know which roll we're
        # currently processing. json_path must always be that same fixed path, regardless
        # of which roll's xlsx_path was passed in. Do NOT derive this from xlsx_path (eg.
        # via .with_suffix('.json')) -- that produces a per-roll path the plugin never
        # looks at, so it silently imports stale/nonexistent data.
        self.json_path = self.script_dir / "metadata.json"
        # TEST ONLY -- see apply_lrplugin(). Pre-created empty right before
        # Importer.lua runs, since Lightroom's process appeared unable to
        # CREATE a brand-new file in this folder earlier (the abandoned
        # signal-file mechanism), only write to an existing one. This checks
        # whether that same workaround lets it write a real data manifest.
        self.vc_manifest_path = self.script_dir / "vc_manifest.txt"
        self.raw_folder = Path(raw_folder) if raw_folder else None
        print(self.xlsx_path)

        db.d("Stage: check excel")

        if not self.xlsx_path.exists():

            if xlsx_path is not None:
                # Caller told us exactly where the roll's xlsx should be; if it's missing
                # that's an upstream error (eg. newRoll.py never ran for this roll), not
                # something to silently recover from by popping a folder picker.
                db.d(f"metadata xlsx not found at {self.xlsx_path}")
                raise FileNotFoundError(f"Expected metadata xlsx at {self.xlsx_path}")

            db.d("metadata.xlsx not found")
            db.d("Stage: select raw folder")

            raw_folder_selected = self.select_raw_folder()
            raw_files = self.get_raw_files_from_folder(raw_folder_selected)

            db.d(f"Raw files found for template: {len(raw_files)}")

            self.generate_template(raw_files=raw_files)

            db.d("Template generated. Populate metadata.xlsx and rerun.")
            sys.exit()

        db.d("Stage: process excel")
        self.data = self.process_excel()

        db.d(f"Exposure rows processed: {len(self.data)}")

        db.d("Stage: export json")
        self.export_json()

        db.d("Stage: import json")
        self.data = self.load_json()

        self.fields = [
            "nlpOriginalCameraMake",
            "nlpOriginalCameraModel",
            "nlpOriginalLensMake",
            "nlpOriginalLens",
            "nlpFilmStock",
            "nlpFilmISO",
            "nlpGearNotes",

            "nlpShotAtIso",
            "nlpAperture",
            "nlpShutterSpeed",
            "nlpFocalLength",
            "nlpDateTaken",
            "nlpShootingNotes",

            "nlpScanEquipment",
            "nlpLightSource",
            "nlpFilmHolder",
            "nlpDigitizationNotes",

            "nlpDeveloper",
            "nlpDevDilution",
            "nlpDevTimeTemp",
            "nlpDevMethod",
            "nlpDevelopmentNotes"
        ]

        self.shared_nlp = self.get_shared_nlp_fields(self.data)
        self.strip_shared_nlp_fields(self.data, self.shared_nlp)

        db.d(f"Shared NLP fields detected: {len(self.shared_nlp)}")
        if self.shared_nlp:
            db.d(f"Shared NLP field names: {list(self.shared_nlp.keys())}")




    def generate_template(self, raw_files=None):

        wb = Workbook()
        ws = wb.active
        ws.title = "Metadata"

        ws.append([
            # File info
            "Index",
            "rawFileName",
            "rawFilePath",
            
            # Date info
            "Year",
            "Month",
            "Day",

            # location info
            "Sublocation",
            "City",
            "State",
            "Country/Region",

            # ID info
            "Intellectual Genre",
            "Scene",

            # camera info
            "Camera Make",
            "Camera Model",
            "Lens Make",
            "Lens Model",

            # Film info
            "Film Stock",
            "Film ISO",
            "Gear Notes",

            # exposure info
            "Shot at ISO",
            "Aperture",
            "Shutter Speed",
            "Focal Length",
            "Shooting Notes",

            # scan info
            "Scan Equipment",
            "Light Source",
            "Film Holder",
            "Digitization Notes",
            
            # dev notes
            "Developer",
            "Dilution",
            "Dev Time/Temp",
            "Dev Method",
            "Dev Notes"
        ])

        if raw_files:
            for i, p in enumerate(raw_files, start=1):
                ws.append([
                    i,
                    p.name,
                    str(p.resolve())
                ])

        wb.save(self.xlsx_path)



    @staticmethod
    def _parse_lens_info(lens_model):
        """Parses '{focal}mm{maxAperture}' (eg. '28mm2.8') out of a free-text Lens
        Model string (eg. '28mm f/2.8'), for appending onto Intellectual Genre
        (the camera(+lens) field -- Scene is the stock field). Mirrors
        exposureObj.py's LensModel-derived maxAperture rules exactly (see
        _update_from_exif() ~lines 355-364) and its cast_lns() 'mm' focal-length
        fallback (~lines 277-281), since that parsing has already been validated
        against real EXIF LensModel strings across the whole archive -- this is
        intentionally not a new/different parser. Returns None if a focal length
        and max aperture can't both be extracted (eg. blank or unexpected text)."""

        if not lens_model:
            return None

        model = str(lens_model).strip()

        focal = None
        if 'mm' in model:
            try:
                focal = int(model.split('mm')[0].strip())
            except ValueError:
                focal = None

        max_aperture = None
        if '/' in model:
            try:
                max_aperture = float(model.split('/')[-1].split(' ')[0])
            except ValueError:
                max_aperture = None
        elif 'mm' in model and 'f' in model:
            try:
                max_aperture = float(model.split('f')[-1].strip())
            except ValueError:
                max_aperture = None

        if focal is None or max_aperture is None:
            return None

        aperture_str = f'{max_aperture:.0f}' if max_aperture > 10 else f'{max_aperture:.1f}'

        return f'{focal}mm{aperture_str}'

    def process_excel(self):

        wb = load_workbook(self.xlsx_path)
        ws = wb.active

        # Seconds offset per row is driven by the Index column's order within
        # each shooting date, NOT by physical row order in the sheet (which
        # tracks rawFileName/scan order). If frames were scanned out of the
        # real shooting sequence, the rows never get physically reordered --
        # only Index is corrected by hand -- so grouping+sorting by Index here
        # is what makes DateTimeOriginal come out chronological.
        date_groups = {}
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            idx, raw, raw_path, year, month, day = row[:6]
            if raw is None or not (year and month and day):
                continue
            key = f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
            date_groups.setdefault(key, []).append((row_num, idx))

        sec_by_row = {}
        for entries in date_groups.values():
            entries.sort(key=lambda e: (e[1] is None, e[1]))
            for sec, (row_num, _idx) in enumerate(entries):
                sec_by_row[row_num] = sec

        data = []
        xlsx_updated = False

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            (
                # File info
                idx,
                raw,
                raw_path,

                # Date info
                year,
                month,
                day,

                # location info
                sublocation,
                city,
                state,
                country,

                # ID info
                intellectual_genre,
                scene,

                # camera info
                cam_make,
                cam_model,
                lens_make,
                lens_model,

                # film info
                film_stock,
                film_iso,
                gear_notes,

                # exposure info
                shot_at_iso,
                aperture,
                shutter,
                focal_length,
                shooting_notes,

                # scan info
                scan_equipment,
                light_source,
                film_holder,
                digitization_notes,

                # dev notes
                developer,
                dev_dilution,
                dev_time_temp,
                dev_method,
                dev_notes

            ) = row[:33]  # tolerate trailing columns (eg. Notes) added by importMetadata.py's Import sheet merge

            if raw is None:
                continue

            # Append the lens's focal length + max aperture to Intellectual
            # Genre (the camera(+lens) field -- Scene is the stock field, eg.
            # (Intellectual Genre="F3", Lens Model="28mm f/2.8") -->
            # Intellectual Genre="F3 28mm2.8". Reuses exposureObj.py's exact
            # LensModel-parsing rules (see exposureObj.py _update_from_exif(),
            # ~lines 355-364, and cast_lns()) rather than a new parser, since
            # that logic is already validated against real EXIF LensModel
            # strings across the whole archive.
            lens_info = self._parse_lens_info(lens_model)
            if lens_info and intellectual_genre not in FIXED_LENS_CAMERA_IDS:
                # Idempotency guard: if this row's already been run through
                # this pass before, Intellectual Genre already ends with (or,
                # if it started blank, already equals) the expected lens
                # suffix -- skip re-appending it, otherwise reruns pile up
                # "F3 28mm2.8 28mm2.8 ...".
                already_applied = intellectual_genre == lens_info or (intellectual_genre and intellectual_genre.endswith(f" {lens_info}"))
                if not already_applied:
                    intellectual_genre = f"{intellectual_genre} {lens_info}" if intellectual_genre else lens_info
                    # Persist into the actual xlsx cell, not just this
                    # in-memory record -- the original ask was for the source
                    # file itself to show "F3 28mm2.8", not only whatever gets
                    # handed to Lightroom. Intellectual Genre is column 11 (K)
                    # -- see METADATA_COLUMNS / generate_template().
                    ws.cell(row=row_num, column=11, value=intellectual_genre)
                    xlsx_updated = True

            date_created = None
            exif_datetime_original = None

            if year and month and day:

                y = int(year)
                m = int(month)
                d = int(day)

                sec = sec_by_row[row_num]

                date_created = f"{y:04d}-{m:02d}-{d:02d}T12:00:{sec:02d}Z"

                # Only prepare EXIF datetime if rawFilePath exists
                if raw_path not in (None, ""):
                    exif_datetime_original = f"{y:04d}:{m:02d}:{d:02d} 12:00:{sec:02d}"

            record = {
                "fileName": raw,
                "rawFilePath": raw_path,

                "standard": {
                    "location": sublocation,
                    "city": city,
                    "stateProvince": state,
                    "country": country,
                    "intellectualGenre": intellectual_genre,
                    "scene": scene,
                    "dateCreated": date_created
                },

                "exif": {
                    "dateTimeOriginal": exif_datetime_original
                },

                "nlp": {
                    "nlpOriginalCameraMake": cam_make,
                    "nlpOriginalCameraModel": cam_model,
                    "nlpOriginalLensMake": lens_make,
                    "nlpOriginalLens": lens_model,

                    "nlpFilmStock": film_stock,
                    "nlpFilmISO": film_iso,
                    "nlpGearNotes": gear_notes,

                    "nlpShotAtIso": shot_at_iso,
                    "nlpAperture": aperture,
                    "nlpShutterSpeed": shutter,
                    "nlpFocalLength": focal_length,
                    "nlpShootingNotes": shooting_notes,

                    "nlpScanEquipment": scan_equipment,
                    "nlpLightSource": light_source,
                    "nlpFilmHolder": film_holder,
                    "nlpDigitizationNotes": digitization_notes,

                    "nlpDeveloper": developer,
                    "nlpDevDilution": dev_dilution,
                    "nlpDevTimeTemp": dev_time_temp,
                    "nlpDevMethod": dev_method,
                    "nlpDevelopmentNotes": dev_notes
                }
            }

            data.append(record)

        if xlsx_updated:
            wb.save(self.xlsx_path)
            db.d(f"Intellectual Genre updated with lens info and saved back to {self.xlsx_path}")

        return data




    def export_json(self):
        # ensure_ascii=False: the default (True) escapes every non-ASCII
        # character as \uXXXX (eg. "…" -> "…") -- valid JSON, but the
        # Lightroom-side plugin's bundled json.lua decoder doesn't
        # implement \uXXXX escapes and throws "Invalid escape sequence" the
        # moment a Shooting/Gear/Dev Notes field (or any free-text field)
        # contains so much as a curly quote or an em dash. Writing the raw
        # UTF-8 characters instead sidesteps that entirely -- Lua strings
        # are just byte sequences, so it reads them fine as-is.
        with open(self.json_path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=4, ensure_ascii=False)

    def load_json(self):

        with open(self.json_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _on_press(self, key):

        if key == keyboard.Key.esc and not self.ignore_esc:
            if not self.stop_flag:
                # Every stage checks stop_flag and returns silently once it's
                # set -- without this print, an ESC press (accidental, or a
                # reaction to what looked like a freeze) makes the whole
                # script just vanish with zero explanation, indistinguishable
                # from an actual crash.
                print("\nESC pressed -- stopping.\n")
            self.stop_flag = True

        if key == keyboard.Key.enter:
            self.enter_event.set()

        if hasattr(key, "char") and key.char == self.acceptButton:
            self.accept_event.set()

    def _sleep(self, seconds, label):
        """time.sleep() wrapper that tallies where fixed delays go, for
        print_diagnostics(). Tallies the requested duration (not measured wall
        time) since the point is to report the *configured* delay values so
        they can be tuned directly -- see print_diagnostics()."""
        if seconds > 0:
            time.sleep(seconds)
        entry = self.timings[label]
        entry[0] += 1
        entry[1] += seconds

    @contextmanager
    def _timed(self, label):
        """Measures actual wall-clock time of a block (menu clicks, subprocess
        calls, human-wait pauses) for print_diagnostics()."""
        t0 = time.perf_counter()
        try:
            yield
        finally:
            entry = self.timings[label]
            entry[0] += 1
            entry[1] += time.perf_counter() - t0

    def press(self, key):

        key_map = {
            "left": Key.left,
            "up": Key.up,
            "down": Key.down,
            "tab": Key.tab,
            "esc": Key.esc,
            "enter": Key.enter,
            "delete": Key.delete,
            "alt": Key.alt,
            "option": Key.alt
        }

        k = key_map.get(key, key)
        press_time = self.delay_keypress

        if key == "esc":
            self.ignore_esc = True
            press_time = 0.2

        self.kb.press(k)
        self._sleep(press_time, "press:hold")
        self.kb.release(k)

        if key == "esc":
            self._sleep(0.05, "press:esc_settle")
            self.ignore_esc = False

        self._sleep(self.delay_default*0, "press:trailing") # TIMESPEED: REDUCED TO ZERO

    def hotkey(self, *keys):

        key_map = {
            "cmd": Key.cmd,
            "ctrl": Key.ctrl,
            "shift": Key.shift,
            "alt": Key.alt,
            "tab": Key.tab,
            "v": "v",
            "a": "a",
            "d": "d",
            "1": "1",
            "enter": Key.enter
        }

        parsed = [key_map.get(k, k) for k in keys]

        for k in parsed[:-1]:
            self.kb.press(k)

        self.kb.press(parsed[-1])
        self.kb.release(parsed[-1])

        for k in reversed(parsed[:-1]):
            self.kb.release(k)

        self._sleep(self.delay_default, "hotkey")

    def paste_text(self, text):

        if text is None or text == "":
            return

        pyperclip.copy(str(text))
        self._sleep(self.delay_paste, "paste_text")
        self.hotkey("cmd", "a")
        self._sleep(self.delay_paste, "paste_text")
        self.hotkey("cmd", "v")
        # Guards the stale-clipboard race, not visible paste settle -- see
        # delay_clipboard_release's comment in __init__. Tracked separately
        # from the "paste_text" label above so diagnostics can distinguish it.
        self._sleep(self.delay_clipboard_release, "paste_text:clipboard_release")

    def calibrate(self):

        # (attribute to set, on-screen prompt). cameraMake_pos anchors the
        # tab-through paste fields (self.fields); the four dropdown fields cannot
        # be reached by tabbing and are selected separately via select_dropdown().
        targets = [
            ("cameraMake_pos", "Camera Make", "CAMERA_MAKE_POS"),
            ("filmFormat_pos", "Film Format", "FILM_FORMAT_POS"),
            ("scanMethod_pos", "Scan Method", "SCAN_METHOD_POS"),
            ("pushPull_pos", "Push-Pull", "PUSH_PULL_POS"),
            ("developedAt_pos", "Developed At", "DEVELOPED_AT_POS"),
        ]

        captured = []

        for attr, label, const_name in targets:

            print("\nCALIBRATION")
            print(f"Move mouse to '{label}' field")
            print(f"Press <{self.acceptButton}> to capture\n")

            self.accept_event.clear()

            while not self.accept_event.is_set():

                if self.stop_flag:
                    return

                time.sleep(0.01)

            pos = pyautogui.position()
            setattr(self, attr, pos)

            print(f"{const_name} = ({pos.x}, {pos.y})")
            captured.append(f"{const_name} = ({pos.x}, {pos.y})")

        print("\nPaste this into the constants block:\n")
        print("\n".join(captured))

    def run_metadata(self, record):

        pyautogui.moveTo(self.cameraMake_pos)
        self._sleep(self.delay_mousemove, "run_metadata:move")
        pyautogui.click()
        self._sleep(self.delay_keypress, "run_metadata:click")

        for field in self.fields:

            if self.stop_flag:
                return

            value = None

            if "nlp" in record and field in record["nlp"]:
                value = record["nlp"][field]

            if value:
                self.paste_text(value)

            self.press("tab")



    def select_dropdown(self, pos, down=0):
        """Selects an option in a Lightroom dropdown/combo-box field that can't be
        reached by tabbing and can't be filled via paste_text() (eg. Film Format,
        Scan Method, Push-Pull, Developed At). Clicks the calibrated position to
        open the dropdown, then ALWAYS presses up DROPDOWN_RESET_PRESSES times
        first, before moving `down` presses to the intended entry, then
        confirming with enter.

        The reset-to-top step matters because a dropdown opens with whatever
        value it currently holds highlighted, not always the first entry -- and
        that value can be blank/empty. A down-count computed assuming "starts at
        entry 1" silently lands on the wrong entry (or nowhere) if the field
        actually started somewhere else. Always resetting first removes that
        assumption entirely, at the cost of a fixed handful of extra keystrokes
        per dropdown -- cheap, since these run once per roll, not per image."""

        pyautogui.moveTo(pos)
        self._sleep(self.delay_mousemove, "select_dropdown:move")
        pyautogui.click()
        self._sleep(self.delay_keypress, "select_dropdown:click")

        # Dropdown/combo-box fields need visibly more time than a text field to
        # register and redraw each arrow-key move -- press()'s normal ~0.004s
        # gap (used everywhere else, eg. tab-through paste fields) is too fast
        # here and keystrokes get dropped/miscounted.
        for _ in range(DROPDOWN_RESET_PRESSES):
            self.press("up")
            self._sleep(self.delay_keypress, "select_dropdown:reset_gap")

        for _ in range(down):
            self.press("down")
            self._sleep(self.delay_keypress, "select_dropdown:down_gap")

        self._sleep(self.delay_dropdown_confirm, "select_dropdown:pre_enter")
        self.press("enter")
        self._sleep(self.delay_dropdown_confirm, "select_dropdown:post_enter")


    def verify_selection_filenames(self):
        """Preflight safeguard, run before ANY field is touched: reads back the
        filenames Lightroom currently has selected (via the read-only "Export
        Selection" Plug-in Extra -- Importer.exportSelection()) and compares
        them, in on-screen order, against self.data's expected fileName
        sequence. run_metadata()/apply_lrplugin() apply rows POSITIONALLY, so a
        stale, wrong, or misordered filmstrip selection would otherwise paste
        frame N's metadata onto the wrong photo with no undo. Returns True and
        proceeds only on an exact match; otherwise prints the mismatch, sets
        stop_flag, and returns False without Lightroom being touched."""

        db.d("Stage: verify Lightroom selection matches expected filenames")

        # select all -- no need to deselect first, cmd+a replaces the prior
        # selection regardless of its state.
        self.hotkey("cmd", "a")
        self._sleep(self.delay_keypress, "verify_selection_filenames:select_all_settle")

        with self._timed("verify_selection_filenames:read_lr_selection"):
            rows = read_lr_selection()
        if rows is None:
            self.stop_flag = True
            return False

        actual = [fname for fname, _path in rows]
        expected = [record["fileName"] for record in self.data]

        if actual != expected:
            print("=" * 70)
            print("ERROR: Lightroom's selected photos don't match the roll's metadata xlsx.")
            print(f"  expected ({len(expected)}): {expected}")
            print(f"  selected ({len(actual)}): {actual}")
            print("Check that the roll is in Quick Collection, sorted by filename")
            print("ascending, with every frame (and only those frames) selected.")
            print("Nothing was written to Lightroom.")
            print("=" * 70)
            self.stop_flag = True
            return False

        db.d(f"Selection verified: {len(actual)} photo(s) match expected order.")
        return True


    def apply_lrplugin(self):

        db.d("Stage: apply Lightroom plugin")

        # select all -- no need to deselect first, cmd+a replaces the prior
        # selection regardless of its state.
        self.hotkey("cmd", "a")
        self._sleep(self.delay_keypress, "apply_lrplugin:select_all_settle")

        # TEST ONLY -- pre-create empty so Importer.lua only ever writes to
        # (never creates) this path. Delete this line once we know whether
        # the write actually succeeds.
        self.vc_manifest_path.write_text("")

        # This is the step that actually applies Sublocation/City/State/Country/
        # Intellectual Genre/Scene/dateCreated (everything in each record's "standard"
        # block) -- it was previously run with stdout/stderr sent to DEVNULL and no
        # returncode check, so a failed menu click (eg. Lightroom not frontmost, or the
        # custom lrplugin-dev plugin not installed/enabled) looked identical to success:
        # nothing printed, script kept going, and none of that metadata ever landed.
        with self._timed("apply_lrplugin:json_import_menu_click"):
            result = click_plugin_extras_menu_item("   JSON Import")

        if result.returncode != 0:
            print("=" * 70)
            print("ERROR: JSON Import menu click failed.")
            print("Sublocation/City/State/Country/Intellectual Genre/Scene/dateCreated")
            print("were NOT applied to any photo. Stopping here rather than continuing")
            print("as if it worked.")
            print(result.stderr.strip())
            print('Check that Lightroom is running and Library > Plug-in Extras >')
            print('"JSON Import" exists in the menu (Plug-in Manager > NLP Importer')
            print('must be enabled).')
            print("=" * 70)
            self.stop_flag = True
            return

        # Runtime here varies (the Lua-side virtual-copy sync scans the whole
        # catalog once per run), so a fixed delay before dismissing the
        # completion dialog isn't reliable, and a file-based "done" signal
        # turned out to be too, for reasons that were never fully pinned
        # down (Lightroom's process silently failed to create a new file in
        # this folder, even though it can read/write existing ones there --
        # see git history if that's ever worth revisiting). Simplest robust
        # option: wait for you to actually press Enter. The same keypress
        # both dismisses Lightroom's dialog (Enter activates its default OK
        # button, since Lightroom is frontmost) and signals this script to
        # continue -- one physical action, no synthetic Enter sent
        # afterward, no guessing about timing.
        print("Waiting for you to press Enter once the 'Import complete' dialog is up...")

        self.enter_event.clear()

        with self._timed("apply_lrplugin:HUMAN_WAIT_for_enter"):
            while not self.enter_event.is_set():
                if self.stop_flag:
                    return
                time.sleep(0.01)

        self.hotkey("cmd", "d")
        self.press("left")


    def select_raw_folder(self):

        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        folder = filedialog.askdirectory(
            title="Select folder containing raw files"
        )

        root.destroy()

        if folder == "":
            return None

        return Path(folder)


    def get_raw_files_from_folder(self, folder_path):

        if folder_path is None:
            return []

        exts = {
            ".arw", ".dng", ".nef", ".cr2", ".cr3", ".raf",
            ".orf", ".rw2", ".pef", ".srw", ".tif", ".tiff"
        }

        files = []

        for p in folder_path.iterdir():
            if p.is_file() and p.suffix.lower() in exts:
                files.append(p)

        files.sort(key=lambda p: p.name.lower())

        return files



    def set_capture_time(file_path, dt_str):
        subprocess.run([
            "exiftool",
            f"-DateTimeOriginal={dt_str}",
            f"-CreateDate={dt_str}",
            f"-ModifyDate={dt_str}",
            "-overwrite_original",
            file_path
        ], check=True)


    def finish_image(self):
        self.press("esc")
        self.press("b")
        self.press("left")
        self._sleep(self.delay_finish_image, "finish_image")


    def save_metadata_sidecars(self):
        db.d("Stage: save metadata sidecars")

        # select all -- no need to deselect first, cmd+a replaces the prior
        # selection regardless of its state.
        self.hotkey("cmd", "a")
        self._sleep(self.delay_keypress, "save_metadata_sidecars:select_all_settle")

        # save metadata to files
        self.hotkey("cmd", "s")
        self._sleep(self.delay_keypress, "save_metadata_sidecars:save_settle")


    def run(self):

        db.d("Stage: run macro")

        print("\nPress ESC anytime to stop\n")

        run_t0 = time.perf_counter()
        try:
            self._run_body()
        finally:
            self.print_diagnostics(time.perf_counter() - run_t0)

    def _run_body(self):

        with self._timed("phase:verify_selection_filenames"):
            ok = self.verify_selection_filenames()
        if not ok:
            return

        # Skip the manual calibration prompts entirely once all five positions
        # are hardcoded above -- only fall back to calibrate() (which re-prompts
        # for and re-captures all five together) if any is still unset.
        if any(p is None for p in (
            self.cameraMake_pos,
            self.filmFormat_pos,
            self.scanMethod_pos,
            self.pushPull_pos,
            self.developedAt_pos,
        )):
            self.calibrate()

        if self.stop_flag:
            return


        with self._timed("phase:save_metadata_sidecars"):
            self.save_metadata_sidecars()

        with self._timed("phase:apply_exif_dates"):
            self.apply_exif_dates()

        if self.stop_flag:
            return

        with self._timed("phase:apply_lrplugin"):
            self.apply_lrplugin()

        if self.stop_flag:
            return

        with self._timed("phase:apply_shared_nlp_metadata"):
            self.apply_shared_nlp_metadata()

        if self.stop_flag:
            return

        with self._timed("phase:apply_dropdown_fields"):
            self.apply_dropdown_fields()

        if self.stop_flag:
            return

        print(f"Starting in {self.delay_start:.1f}s...")
        self._sleep(self.delay_start*2, "run:start_delay")

        idx = 0

        with self._timed("phase:per_image_loop"):
            while idx < len(self.data):

                if self.stop_flag:
                    break

                record = self.data[idx]

                # Shared/roll-wide nlp fields were already stripped out of every record by
                # strip_shared_nlp_fields() (and already applied once via
                # apply_shared_nlp_metadata()). If nothing unique is left for this image,
                # skip the whole click-through-every-field-and-tab sequence in run_metadata()
                # -- that's the per-image cost that was slowing this down for fields that are
                # roll-wide statics anyway. Still call finish_image() so the filmstrip
                # position advances and stays in sync with self.data.
                nlp_block = record.get("nlp", {})
                has_unique_data = any(v not in (None, "") for v in nlp_block.values())

                if has_unique_data:
                    self.run_metadata(record)

                    if self.stop_flag:
                        break
                else:
                    db.d(f"Skip image {idx + 1}/{len(self.data)}: no unique (non-shared) metadata")

                self.finish_image()

                idx += 1
                self.images_processed = idx

                db.d(f"Processed exposure {idx}/{len(self.data)}")

        print("Finished JSON records")
        time.sleep(0.5)  # give the last finish_image()'s ESC+click time to settle before the final print

    def print_diagnostics(self, total_elapsed):
        """Prints a breakdown of where run()'s wall-clock time actually went --
        phases (self._timed() blocks: menu clicks, subprocess calls, the
        per-image loop, the human "click Enter" wait) and fixed delays
        (self._sleep() calls, tallied by the requested duration of each). Run
        with different self.delay_* values / DROPDOWN_RESET_PRESSES and diff
        the two tables to see what actually moved."""

        human_labels = {label for label in self.timings if "HUMAN_WAIT" in label}
        phase_labels = {label for label in self.timings if label.startswith("phase:")}
        sleep_labels = set(self.timings) - human_labels - phase_labels

        human_total = sum(self.timings[l][1] for l in human_labels)
        sleep_total = sum(self.timings[l][1] for l in sleep_labels)

        print("\n" + "=" * 70)
        print("TIMING DIAGNOSTICS")
        print("=" * 70)
        print(f"Total wall-clock:        {total_elapsed:6.1f}s")
        print(f"Images processed:        {self.images_processed}"
              + (f"  ({total_elapsed / self.images_processed:.2f}s/image overall)"
                 if self.images_processed else ""))
        print(f"Human wait (not tunable):{human_total:6.1f}s")
        print(f"Tracked fixed delays:    {sleep_total:6.1f}s  "
              f"({sleep_total / total_elapsed * 100:.0f}% of total)" if total_elapsed else "")

        if phase_labels:
            print("\n-- Phases (self._timed blocks) --")
            for label in sorted(phase_labels, key=lambda l: -self.timings[l][1]):
                count, total = self.timings[label]
                print(f"  {label:<45} {total:7.2f}s  ({count} call(s), avg {total / count:.3f}s)")

        if human_labels:
            print("\n-- Human-wait pauses (excluded from 'tracked fixed delays' -- not tunable) --")
            for label in sorted(human_labels, key=lambda l: -self.timings[l][1]):
                count, total = self.timings[label]
                print(f"  {label:<45} {total:7.2f}s  ({count} call(s))")

        if sleep_labels:
            print("\n-- Fixed delays, by call site (tunable -- see self.delay_* / literals) --")
            for label in sorted(sleep_labels, key=lambda l: -self.timings[l][1]):
                count, total = self.timings[label]
                print(f"  {label:<45} {total:7.2f}s  ({count} call(s), avg {total / count:.4f}s)")

        print("=" * 70 + "\n")


    def apply_exif_dates(self):
        """Batched replacement for the old one-exiftool-process-per-image loop
        -- same net effect (check each raw file's current DateTimeOriginal,
        write corrected values only where they differ, refresh Lightroom if
        anything changed), but the check and the write are each a single
        exiftool invocation covering every image, not N invocations. Process
        spawn overhead dominated the old per-file cost (see
        apply_exif_dates:exiftool_check/:exiftool_write in early diagnostics
        runs), so batching beats parallelizing N single-file calls -- it cuts
        the number of processes started instead of just overlapping them."""

        db.d("Stage: apply EXIF DateTimeOriginal")

        candidates = []  # (i, raw_path, dt_original)
        for i, record in enumerate(self.data, start=1):
            raw_path = record.get("rawFilePath")
            exif_block = record.get("exif", {})
            dt_original = exif_block.get("dateTimeOriginal")

            if not raw_path or not dt_original:
                db.d(f"Skip EXIF date {i}: missing rawFilePath or dateTimeOriginal")
                continue

            candidates.append((i, raw_path, dt_original))

        if not candidates:
            db.d("No EXIF date changes needed; Lightroom refresh skipped")
            return

        # Batch read: mirrors rollObj.py's fetch_exif() -- one JSON call for
        # every candidate, correlated back to source files by "SourceFile"
        # rather than by output order, since exiftool silently DROPS any file
        # it couldn't read from the JSON array instead of erroring the whole
        # batch (confirmed: a missing/corrupt file just isn't in the result).
        raw_paths = [raw_path for _, raw_path, _ in candidates]
        with self._timed("apply_exif_dates:exiftool_check_batch"):
            check_result = subprocess.run(
                ["exiftool", "-j", "-fast2", "-DateTimeOriginal", *raw_paths],
                capture_output=True, text=True
            )

        try:
            current_by_path = {
                rec["SourceFile"]: rec.get("DateTimeOriginal", "")
                for rec in json.loads(check_result.stdout or "[]")
            }
        except json.JSONDecodeError:
            db.d(f"ExifTool batch read produced no usable JSON: {check_result.stderr.strip()}")
            current_by_path = {}

        to_write = []  # (i, raw_path, xmp_path, dt_original)
        for i, raw_path, dt_original in candidates:
            if raw_path not in current_by_path:
                db.d(f"ExifTool read error on {raw_path}: not returned by batch read ({check_result.stderr.strip()})")
                continue

            if current_by_path[raw_path] == dt_original:
                db.d(f"Skip EXIF date {i}/{len(self.data)}: already correct")
                continue

            to_write.append((i, raw_path, Path(raw_path).with_suffix(".xmp"), dt_original))

        if not to_write:
            db.d("No EXIF date changes needed; Lightroom refresh skipped")
            return

        # Batch write: one exiftool call for every sidecar that needs a new
        # date, via a CSV of per-file values -- confirmed exiftool applies
        # each row's tags to its own SourceFile within a single invocation
        # (different files can get different values in the same call). The
        # CSV's SourceFile column must match the path string handed to
        # exiftool on the command line exactly, or the row won't match.
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="") as f:
            csv_path = f.name
            writer = csv.writer(f)
            writer.writerow(["SourceFile", "DateTimeOriginal", "CreateDate", "ModifyDate"])
            for _, _, xmp_path, dt_original in to_write:
                writer.writerow([str(xmp_path), dt_original, dt_original, dt_original])

        try:
            with self._timed("apply_exif_dates:exiftool_write_batch"):
                write_result = subprocess.run(
                    ["exiftool", f"-csv={csv_path}", "-overwrite_original",
                     *[str(xmp_path) for _, _, xmp_path, _ in to_write]],
                    capture_output=True, text=True
                )
        finally:
            os.unlink(csv_path)

        # exiftool's returncode is nonzero if ANY file in the batch failed,
        # even when most succeeded -- so per-file success/failure is read from
        # stderr's "Error: ... - <path>" lines, not from the returncode alone
        # (confirmed: files with no error still get written in the same run).
        failed_paths = set()
        for line in write_result.stderr.splitlines():
            if line.startswith("Error:") and " - " in line:
                failed_paths.add(line.rsplit(" - ", 1)[-1].strip())

        changed_any = False
        for i, raw_path, xmp_path, _ in to_write:
            if str(xmp_path) in failed_paths:
                db.d(f"ExifTool write error on {raw_path}: see stderr above ({xmp_path})")
            else:
                db.d(f"EXIF date set {i}/{len(self.data)}: {raw_path}")
                changed_any = True

        if changed_any:
            self.refresh_lr_metadata_from_files()
        else:
            db.d("No EXIF date changes needed; Lightroom refresh skipped")



    def refresh_lr_metadata_from_files(self):

        db.d("Stage: Lightroom save/read metadata from files")

        # select all -- no need to deselect first, cmd+a replaces the prior
        # selection regardless of its state.
        self.hotkey("cmd", "a")
        self._sleep(self.delay_keypress, "refresh_lr_metadata_from_files:select_all_settle")

        db.d("Stage: refresh -- waiting for selection to settle")
        self._sleep(self.delay_keypress, "refresh_lr_metadata_from_files:settle")

        # Menu item title: "Read Metadata from Files" (plural) -- Lightroom relabels this
        # command based on selection count (singular "...from File" for a single photo).
        # Our workflow always selects the whole roll, so plural is correct here.
        script = '''
        tell application "Adobe Lightroom Classic" to activate
        delay 0.3
        tell application "System Events"
            tell process "Adobe Lightroom Classic"
                click menu bar item "Metadata" of menu bar 1
                delay 0.3
                tell menu item "Read Metadata from Files" of menu 1 of menu bar item "Metadata" of menu bar 1
                    perform action "AXPress"
                end tell
                delay 0.3
            end tell
        end tell
        '''

        db.d("Stage: refresh -- running AppleScript menu click")

        with self._timed("refresh_lr_metadata_from_files:menu_click"):
            result = subprocess.run(
                ["osascript", "-e", script],
                capture_output=True,
                text=True
            )

        db.d("Stage: refresh -- menu click done")

        if result.returncode != 0:
            # Non-fatal: apply_exif_dates() already wrote the corrected date into the
            # .xmp sidecar on disk regardless of whether Lightroom's UI picked it up here,
            # so nothing is lost -- Lightroom just won't show the updated date until it
            # re-reads that file some other way (eg. manually, or next relaunch).
            print("WARNING: 'Read Metadata from Files' menu click failed -- Lightroom's")
            print("displayed date may be stale, but the .xmp sidecar on disk is correct.")
            print(result.stderr.strip())

        db.d("Stage: refresh -- done")


    def strip_shared_nlp_fields(self, data, shared_nlp):

        if not shared_nlp:
            return

        for record in data:
            nlp_block = record.get("nlp", {})

            for field in shared_nlp:
                if field in nlp_block:
                    del nlp_block[field]


    def apply_shared_nlp_metadata(self):

        if not self.shared_nlp:
            db.d("Stage: apply shared NLP metadata skipped (none detected)")
            return

        db.d("Stage: apply shared NLP metadata")

        # select all -- no need to deselect first, cmd+a replaces the prior
        # selection regardless of its state.
        self.hotkey("cmd", "a")
        self._sleep(self.delay_keypress, "apply_shared_nlp_metadata:select_all_settle")

        # open metadata panel at calibrated field
        pyautogui.moveTo(self.cameraMake_pos)
        self._sleep(self.delay_mousemove, "apply_shared_nlp_metadata:move")
        pyautogui.click()
        self._sleep(self.delay_keypress, "apply_shared_nlp_metadata:click")

        for field in self.fields:

            if self.stop_flag:
                return

            value = self.shared_nlp.get(field)

            if value not in (None, ""):
                self.paste_text(value)

            self.press("tab")

        # close metadata editing (exit field focus) -- stay on the full-roll
        # selection here rather than deselecting down to a single photo.
        # apply_dropdown_fields() runs next and also needs every photo
        # selected, so dropping to one photo and reselecting all a few lines
        # later was pure wasted deselect/reselect.
        self.press("esc")
        self._sleep(self.delay_field_close, "apply_shared_nlp_metadata:close_delay")


    # Dropdown entries in on-screen order, index 0 = first item (the item a
    # freshly-opened dropdown highlights by default). select_dropdown()'s "down"
    # count for a target entry is its index here.
    FILM_FORMAT_ORDER = ["half-frame", "35mm", "645", "6x6", "6x7", "6x8", "6x9"]

    # cameralist.xlsx's filmformat column doesn't use the dropdown's own labels for
    # every value -- 135-format cameras are stored as the (int or str) spool format
    # "135" rather than "35mm", and the sheet has no way to flag half-frame vs.
    # full-frame within that "135" value (no half-frame camera exists in the sheet
    # today). Confirmed against the actual data: every 135-camera row's filmformat
    # is literally 135; 120-camera rows already use the frame-geometry strings
    # (6x6/6x7) that match the dropdown directly, so only "135" needs aliasing.
    FILM_FORMAT_ALIASES = {"135": "35mm"}

    @staticmethod
    def _normalize_format(s):
        # Tolerant match against cameralist.xlsx's filmformat column -- eg. "half
        # frame" vs "half-frame" -- so the sheet's text doesn't need to be
        # byte-for-byte identical to the Lightroom dropdown's labels.
        return "".join(str(s).lower().split()).replace("-", "")

    def _load_cameralist(self):
        # data/ is gitignored and lives at the repo root, one level above
        # lrplugin-dev/ (where this file lives).
        cameralist_path = self.script_dir.parent / "data" / "cameralist.xlsx"

        if not cameralist_path.exists():
            print(f"WARNING: cameralist.xlsx not found at {cameralist_path} -- Film Format cannot be auto-selected.")
            return {}

        wb = load_workbook(cameralist_path)
        ws = wb.active

        header = [c.value for c in ws[1]]
        col = {name: i for i, name in enumerate(header) if name is not None}

        required = ("brand", "model", "filmformat")
        missing = [c for c in required if c not in col]
        if missing:
            print(f"WARNING: cameralist.xlsx missing column(s) {missing} -- Film Format cannot be auto-selected.")
            return {}

        lookup = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            brand = row[col["brand"]]
            model = row[col["model"]]
            filmformat = row[col["filmformat"]]
            if not brand or not model or not filmformat:
                continue
            key = (self._normalize_format(brand), self._normalize_format(model))
            lookup[key] = filmformat

        return lookup

    def _resolve_filmformat_down_count(self):
        """Looks up the roll's camera in cameralist.xlsx and returns the down-arrow
        count for select_dropdown() to land on the matching Film Format entry
        (dropdown highlights the first item by default, so count == list index).
        Returns None -- callers must then skip the Film Format step -- if the
        camera or its filmformat can't be resolved; guessing here means silently
        writing the wrong format into Lightroom with no undo."""

        cam_make = self.shared_nlp.get("nlpOriginalCameraMake")
        cam_model = self.shared_nlp.get("nlpOriginalCameraModel")

        if not cam_make or not cam_model:
            # Camera wasn't flagged as constant across the whole roll (or is
            # blank in the first row) -- fall back to the first record's value.
            first_nlp = self.data[0].get("nlp", {}) if self.data else {}
            cam_make = cam_make or first_nlp.get("nlpOriginalCameraMake")
            cam_model = cam_model or first_nlp.get("nlpOriginalCameraModel")

        if not cam_make or not cam_model:
            print("WARNING: no Camera Make/Model available -- skipping Film Format selection.")
            return None

        cameralist = self._load_cameralist()
        key = (self._normalize_format(cam_make), self._normalize_format(cam_model))
        filmformat = cameralist.get(key)

        if not filmformat:
            print(f"WARNING: '{cam_make} {cam_model}' not found in cameralist.xlsx -- skipping Film Format selection.")
            return None

        norm_order = [self._normalize_format(f) for f in self.FILM_FORMAT_ORDER]

        raw_key = str(filmformat).strip()
        aliased = self.FILM_FORMAT_ALIASES.get(raw_key, filmformat)
        norm_format = self._normalize_format(aliased)

        if norm_format not in norm_order:
            print(f"WARNING: cameralist.xlsx filmformat '{filmformat}' not in known list {self.FILM_FORMAT_ORDER} -- skipping Film Format selection.")
            return None

        return norm_order.index(norm_format)


    def apply_dropdown_fields(self):
        """Selects Film Format / Scan Method / Push-Pull / Developed At once for
        the whole roll. These are Lightroom combo-box fields, not free text, so
        they can't go through paste_text()/tab like self.fields -- and in
        practice they're constant for an entire roll (same camera, same scan
        session, same lab), so -- like apply_shared_nlp_metadata() -- this runs
        once with every photo selected rather than once per frame."""

        if not any([self.filmFormat_pos, self.scanMethod_pos, self.pushPull_pos, self.developedAt_pos]):
            db.d("Stage: apply dropdown fields skipped (not calibrated)")
            return

        db.d("Stage: apply dropdown fields (Film Format / Scan Method / Push-Pull / Developed At)")

        # select all -- no need to deselect first: cmd+a selects everything
        # regardless of prior selection state, so the deselect+redraw that used
        # to run here first was pure wasted time.
        self.hotkey("cmd", "a")
        self._sleep(self.delay_keypress, "apply_dropdown_fields:select_all_settle")

        # Film Format -- data-driven from cameralist.xlsx, skipped if unresolvable.
        if self.filmFormat_pos is not None:
            down_count = self._resolve_filmformat_down_count()
            if down_count is not None:
                self.select_dropdown(self.filmFormat_pos, down=down_count)

        # Scan Method -- hardcoded, always the default/current (first) entry.
        if self.scanMethod_pos is not None:
            self.select_dropdown(self.scanMethod_pos)

        # Push-Pull -- hardcoded, always the 4th entry.
        if self.pushPull_pos is not None:
            self.select_dropdown(self.pushPull_pos, down=3)

        # Developed At -- hardcoded, always the 2nd entry.
        if self.developedAt_pos is not None:
            self.select_dropdown(self.developedAt_pos, down=1)

        # close metadata editing and return to single-image workflow -- "up"
        # collapses the roll-wide selection back down to a single image in one
        # step (see delay_dropdown_deselect_settle's comment in __init__).
        self.press("up")
        self._sleep(self.delay_dropdown_deselect_settle, "apply_dropdown_fields:deselect_settle")


    def get_shared_nlp_fields(self, data):

        shared = {}

        if not data:
            return shared

        for field in self.fields:

            first_value = data[0].get("nlp", {}).get(field)

            if first_value in (None, ""):
                continue

            same_for_all = True

            for record in data[1:]:
                value = record.get("nlp", {}).get(field)

                if value != first_value:
                    same_for_all = False
                    break

            if same_for_all:
                shared[field] = first_value

        return shared


if __name__ == "__main__":
    t1 = time.time()
    tool = metadataTool()
    tool.pause_field = False
    tool.pause_nextImage = False
    tool.run()
    t2 = time.time()

    print(f'Completed in {t2-t1:.2f}s')