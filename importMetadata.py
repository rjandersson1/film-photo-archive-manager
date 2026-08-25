# importMetadata.py
#
# Applies per-frame metadata from a roll's {roll}_metadata.xlsx (built by newRoll.py)
# into Lightroom, via lrplugin-dev/metadataTool.py's existing GUI-automation pipeline.
#
# Why this exists: metadataTool.py applies rows POSITIONALLY -- it tabs through whatever
# photos are currently selected in Lightroom's filmstrip, in order, pasting row 1's data
# onto the 1st selected photo, row 2's onto the 2nd, etc. If you delete bad scans/dupes
# during editing after newRoll.py already generated the xlsx, the row count and the actual
# raw file count drift apart, and the automation will silently paste the WRONG frame's
# metadata onto a photo. There's no undo for that once it's written into Lightroom, so this
# refuses to run the automation on any mismatch, and offers to auto-fix the (disposable,
# easily regenerated) xlsx instead -- Lightroom itself is never touched until the two are
# confirmed to match.
#
# Workflow:
#   1) pick the roll folder to work on -- auto-detected from whatever's currently
#      selected in Lightroom (see detect_roll_from_lr_selection()) if that resolves to
#      exactly one roll under LIBRARY_PATH; otherwise falls back to prompting for a roll
#      index (same LIBRARY_PATH/index convention as newRoll.py), or -1 to browse for it
#   2) SAFEGUARD: reconcile the xlsx's rawFileName rows against what's actually still in
#      01_scans right now
#   3) hand off to lrplugin-dev/metadataTool.py for the actual Lightroom automation, which
#      re-verifies the roll's expected filenames against whatever's actually selected in
#      Lightroom before writing anything (see metadataTool.verify_selection_filenames())
#
# Usage:
#   python importMetadata.py

import os
import time
import subprocess
import importlib.util

from openpyxl import load_workbook
from tkinter import Tk, filedialog

import collectionObj
from newRoll import (
    LIBRARY_PATH, RAW_EXTS, METADATA_COLUMNS, IMPORT_COLUMNS,
    list_raw_files, force_text_format, build_import_sheet,
)


def activate_lightroom():
    # Matches syncVCs.py's activate_lightroom() -- ensures Lightroom is
    # frontmost before any Lightroom-focused automation runs, regardless of
    # whether the roll was auto-detected from the current LR selection or
    # typed at the terminal.
    subprocess.run(
        ["osascript", "-e", 'tell application "Adobe Lightroom Classic" to activate'],
        capture_output=True,
        text=True
    )
    time.sleep(0.3)


def find_roll_folder(library_path, index):
    index_str = str(index).zfill(3)
    matches = []
    if os.path.isdir(library_path):
        for name in sorted(os.listdir(library_path)):
            if name.startswith('.'):
                continue
            full = os.path.join(library_path, name)
            if not os.path.isdir(full):
                continue
            token = name.split('_')[0] if '_' in name else name.split(' - ')[0]
            token = token.strip()
            if token == index_str or (token.isdigit() and int(token) == index):
                matches.append(full)
    return matches


def browse_for_roll_folder():
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    folder = filedialog.askdirectory(
        initialdir=LIBRARY_PATH,
        title="Select roll folder"
    )

    root.destroy()

    if folder == "":
        return None

    return folder


def roll_index_from_folder_name(roll_root):
    # Same "{index}_..." token convention as find_roll_folder() -- best-effort
    # only, since a folder picked by hand isn't guaranteed to follow it.
    name = os.path.basename(roll_root.rstrip('/'))
    token = name.split('_')[0] if '_' in name else name.split(' - ')[0]
    token = token.strip()
    return int(token) if token.isdigit() else None


def prompt_roll_folder():
    while True:
        raw = input('Roll index (-1 to browse): ').strip()
        if raw == '-1':
            folder = browse_for_roll_folder()
            if folder is None:
                print('No folder selected.')
                continue
            return folder, roll_index_from_folder_name(folder)
        if not raw.isdigit():
            print('Enter a numeric roll index, or -1 to browse.')
            continue
        index = int(raw)
        matches = find_roll_folder(LIBRARY_PATH, index)
        if len(matches) == 1:
            return matches[0], index
        if len(matches) == 0:
            print(f'No roll folder found for index {index} in {LIBRARY_PATH}')
            continue
        print(f'Multiple folders matched index {index}:')
        for m in matches:
            print(f'  {m}')
        print('Rename/clean up so only one folder matches, then retry.')


def find_metadata_xlsx(roll_root):
    # Excel/Numbers drop a hidden lock file (eg. "~$foo_metadata.xlsx") next to any
    # xlsx that's currently open on Mac -- ignore those, they're not real data files.
    candidates = [
        f for f in os.listdir(roll_root)
        if f.lower().endswith('_metadata.xlsx') and not f.startswith('~$')
    ]
    if len(candidates) == 0:
        return None
    if len(candidates) > 1:
        print(f'Multiple *_metadata.xlsx files found in {roll_root}: {candidates}')
        print('Leave only one and retry.')
        return None
    return os.path.join(roll_root, candidates[0])


# -------- Safeguard: reconcile xlsx rows against what's actually on disk now --------

def read_xlsx_rows(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb['Metadata']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None:  # rawFileName
            continue
        rows.append(row)
    return wb, ws, rows


def reconcile(xlsx_path, scans_path):
    _, _, rows = read_xlsx_rows(xlsx_path)
    xlsx_files = {row[1] for row in rows}
    disk_files = set(list_raw_files(scans_path))

    missing_on_disk = sorted(xlsx_files - disk_files)   # rows referencing deleted/renamed raws
    missing_in_xlsx = sorted(disk_files - xlsx_files)   # raws with no row (rescans, renames)

    return missing_on_disk, missing_in_xlsx, len(xlsx_files), len(disk_files)


def parse_roll_tokens(roll_root):
    # folder name convention (see newRoll.py): {index}_{date}_{STK}_{CAM}_{Name...}
    name = os.path.basename(roll_root.rstrip('/'))
    parts = name.split('_')
    stk = parts[2] if len(parts) > 2 else None
    cam = parts[3] if len(parts) > 3 else None
    return stk, cam


def auto_fix_xlsx(xlsx_path, scans_path, roll_root, collection):
    """Drops rows whose raw file no longer exists, and appends rows for any raw files
    missing a row -- prefilled with Camera Make/Model + Film Stock/ISO the same way
    newRoll.py does, using the STK/CAM tokens parsed back out of the roll folder name.
    Never touches Lightroom -- this only rewrites the (disposable) xlsx."""
    wb, ws, rows = read_xlsx_rows(xlsx_path)
    disk_files = set(list_raw_files(scans_path))

    kept = [row for row in rows if row[1] in disk_files]
    kept_names = {row[1] for row in kept}
    new_names = sorted(disk_files - kept_names)

    stk_code, cam_id = parse_roll_tokens(roll_root)
    stk_entry = collection.stocklist.get(stk_code, {})
    cam_entry = {}
    for entry in collection.cameralist.values():
        if entry.get('id') == cam_id:
            cam_entry = entry
            break

    header = [cell.value for cell in ws[1]] or METADATA_COLUMNS

    # Rebuild the Metadata sheet WITHIN the same, already-open workbook
    # (not a brand-new Workbook()) so the Import sheet isn't silently
    # dropped by this rewrite.
    new_ws = wb.create_sheet('Metadata_rebuilt')
    new_ws.append(header)

    combined = list(kept) + [
        (None, name, os.path.join(scans_path, name)) + (None,) * (len(header) - 3)
        for name in new_names
    ]
    # sort by rawFileName so row order matches how the files will sort in Lightroom's grid
    combined.sort(key=lambda r: r[1])

    for i, row in enumerate(combined, start=1):
        row = list(row)
        row[0] = i
        if row[1] in new_names:
            if len(row) > 12: row[12] = cam_entry.get('brand')   # Camera Make
            if len(row) > 13: row[13] = cam_entry.get('model')   # Camera Model
            if len(row) > 16: row[16] = stk_entry.get('stock')   # Film Stock
            if len(row) > 17: row[17] = stk_entry.get('boxspeed')  # Film ISO
        new_ws.append(row)

    del wb['Metadata']
    new_ws.title = 'Metadata'
    force_text_format(new_ws, IMPORT_COLUMNS)
    wb.active = wb.sheetnames.index('Metadata')
    wb.save(xlsx_path)
    print(f'Rewrote {xlsx_path}:')
    print(f'  kept:    {len(kept)}')
    print(f'  dropped: {len(rows) - len(kept)} stale row(s) (raw no longer on disk)')
    print(f'  added:   {len(new_names)} new row(s) (raw on disk with no row)')


def ensure_workbook_structure(xlsx_path):
    """Backward-compat migration for roll xlsx files that predate the Import
    sheet -- adds whatever's missing (Import sheet, text formatting).
    Idempotent: safe to call on every run, a no-op once a file is already
    migrated. Notes is Import-only and never added to Metadata -- see
    merge_import_into_metadata()."""
    wb = load_workbook(xlsx_path)

    if 'Metadata' not in wb.sheetnames:
        wb.active.title = 'Metadata'   # every writer in this repo already names it this; defensive fallback

    ws = wb['Metadata']

    if 'Import' not in wb.sheetnames:
        rows = [
            (row[0], row[1])
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[1] is not None
        ]
        build_import_sheet(wb, rows)

    force_text_format(ws, IMPORT_COLUMNS)
    wb.active = wb.sheetnames.index('Metadata')
    wb.save(xlsx_path)


# -------- Import sheet -> Metadata sheet merge --------

def _is_blank(value):
    return value is None or (isinstance(value, str) and value.strip() == '')


def normalize_shutter_speed_for_import(value):
    """Rules for a user-typed Shutter Speed cell on the Import sheet:
      - blank -> nothing to merge
      - bare integer (eg. 250, or the string "250" -- the Import sheet's
        columns are forced to Text format, so Excel stores a typed number
        as a literal digit string, not a Python int/float; both are
        treated the same here) -> reciprocal fraction, "1/250" (shorthand
        for sub-1s speeds)
      - any other string (eg. "1/125", "5s", "1h30m", "1h43m20s") -> kept
        exactly as typed, already unambiguous
      - a plain decimal with no unit suffix (eg. 3.5, or the string "3.5")
        -> ambiguous (could mean 3.5 seconds, could be a mistyped
        fraction) -- warn and skip rather than guess
    """
    if _is_blank(value):
        return None
    if isinstance(value, bool):  # bool is an int subclass -- guard before the int check
        return None
    if isinstance(value, int):
        return f'1/{value}'
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.isdigit():
            return f'1/{stripped}'
        try:
            float(stripped)
        except ValueError:
            return value   # not a plain number -- already formatted ("5s", "1/125", "1h30m", ...)
        print(f'Warning: ambiguous Shutter Speed value {value!r} in Import sheet '
              f'(a plain decimal with no unit) -- leaving blank. Type it as eg. '
              f'"{stripped}s" if that\'s seconds, or as a fraction like "1/{stripped}".')
        return None
    if isinstance(value, float):
        print(f'Warning: ambiguous Shutter Speed value {value!r} in Import sheet '
              f'(Excel likely auto-converted a typed value) -- leaving blank. '
              f'Retype it with the cell set to Text format.')
        return None
    return None


def infer_focal_length(lens_make, lens_model, focal_length_value, lenslist):
    """Only fires when Focal Length is blank and both Lens Make/Model are
    filled -- looks up "{make} {model}" (normalized) in data/lenslist.xlsx."""
    if not _is_blank(focal_length_value):
        return focal_length_value
    if _is_blank(lens_make) or _is_blank(lens_model):
        return focal_length_value
    key = f'{lens_make} {lens_model}'.strip().lower()
    entry = lenslist.get(key)
    if entry and entry.get('focalLength'):
        return entry['focalLength']
    return focal_length_value


def infer_lens_model(lens_make, focal_length, lens_model_value, format_hint, lenslist_by_make_focal):
    """Reverse of infer_focal_length: only fires when Lens Model is blank
    and both Lens Make + Focal Length are filled. A make+focal-length pair
    isn't always unique on its own (eg. a 135-format and a 6x7-format lens
    can share the same nominal focal length) -- format_hint (the roll's
    camera's filmformat, from cameralist.xlsx) disambiguates when more than
    one candidate matches. Never guesses if it's still ambiguous after
    that -- leaves blank and warns rather than picking wrong."""
    if not _is_blank(lens_model_value):
        return lens_model_value
    if _is_blank(lens_make) or _is_blank(focal_length):
        return lens_model_value

    key = (str(lens_make).strip().lower(), str(focal_length).strip())
    candidates = lenslist_by_make_focal.get(key, [])
    if not candidates:
        return lens_model_value

    if len(candidates) > 1 and format_hint:
        filtered = [c for c in candidates if c['format'].strip().lower() == format_hint.strip().lower()]
        if filtered:
            candidates = filtered

    if len(candidates) == 1:
        return candidates[0]['model']

    names = ', '.join(sorted({c['model'] for c in candidates}))
    print(f'Warning: "{lens_make} {focal_length}mm" matches {len(candidates)} lenses in '
          f'lenslist.xlsx ({names}) -- add/fix the "format" column there to disambiguate, '
          f'or fill in Lens Model by hand. Leaving blank.')
    return lens_model_value


def prompt_resolve_conflict(raw_file_name, column_name, a_value, b_value):
    print(f'\nConflict on "{raw_file_name}" / {column_name}:')
    print(f'  [a] Import:   {a_value!r}')
    print(f'  [b] Metadata: {b_value!r}')
    while True:
        raw = input('  Keep which? [a/B] ').strip().lower()
        if raw in ('', 'b'):
            return b_value
        if raw == 'a':
            return a_value
        print('  Enter "a" or "b" (blank keeps Metadata\'s existing value).')


def merge_import_into_metadata(xlsx_path, collection):
    """Merges the Import sheet's user-filled values into Metadata, row-
    matched by rawFileName. Only-one-side-filled cells merge silently;
    both-filled-and-different cells stop the silent merge and prompt
    interactively per cell (prompt_resolve_conflict) rather than either
    guessing or aborting the whole run."""
    wb = load_workbook(xlsx_path)
    if 'Import' not in wb.sheetnames:
        return
    ws_a = wb['Import']
    ws_b = wb['Metadata']

    header_a = [c.value for c in ws_a[1]]
    header_b = [c.value for c in ws_b[1]]
    col_idx_b = {name: i + 1 for i, name in enumerate(header_b) if name is not None}

    name_col_b = col_idx_b.get('rawFileName')
    if name_col_b is None:
        return

    cam_make_col_b = col_idx_b.get('Camera Make')
    cam_model_col_b = col_idx_b.get('Camera Model')

    rows_b_by_name = {}
    for r in range(2, ws_b.max_row + 1):
        fname = ws_b.cell(row=r, column=name_col_b).value
        if fname is not None:
            rows_b_by_name[fname] = r

    # Notes is Import-only and never merges into Metadata (it's a personal
    # scratch field, not part of the schema handed to Lightroom).
    shared_columns = [c for c in IMPORT_COLUMNS if c not in ('rawFileName', 'Notes')]
    merged_count = 0
    conflict_count = 0

    for r in range(2, ws_a.max_row + 1):
        row_a = {header_a[i]: ws_a.cell(row=r, column=i + 1).value
                 for i in range(len(header_a)) if header_a[i] is not None}
        fname = row_a.get('rawFileName')
        if _is_blank(fname):
            continue
        if fname not in rows_b_by_name:
            print(f'Warning: "{fname}" is in Import but has no matching Metadata row -- skipping.')
            continue
        b_row = rows_b_by_name[fname]

        format_hint = None
        if cam_make_col_b and cam_model_col_b:
            cam_make = ws_b.cell(row=b_row, column=cam_make_col_b).value
            cam_model = ws_b.cell(row=b_row, column=cam_model_col_b).value
            if not _is_blank(cam_make) and not _is_blank(cam_model):
                cam_entry = collection.cameralist.get(f'{cam_make} {cam_model}'.strip())
                if cam_entry:
                    format_hint = cam_entry.get('filmformat')

        row_a['Shutter Speed'] = normalize_shutter_speed_for_import(row_a.get('Shutter Speed'))
        row_a['Focal Length'] = infer_focal_length(
            row_a.get('Lens Make'), row_a.get('Lens Model'),
            row_a.get('Focal Length'), collection.lenslist,
        )
        row_a['Lens Model'] = infer_lens_model(
            row_a.get('Lens Make'), row_a.get('Focal Length'),
            row_a.get('Lens Model'), format_hint, collection.lenslist_by_make_focal,
        )

        for col in shared_columns:
            a_val = row_a.get(col)
            if _is_blank(a_val) or col not in col_idx_b:
                continue
            b_cell = ws_b.cell(row=b_row, column=col_idx_b[col])
            b_val = b_cell.value
            if _is_blank(b_val):
                b_cell.value = a_val
                merged_count += 1
            elif str(a_val).strip() != str(b_val).strip():
                resolved = prompt_resolve_conflict(fname, col, a_val, b_val)
                if resolved != b_val:
                    b_cell.value = resolved
                conflict_count += 1

    wb.active = wb.sheetnames.index('Metadata')
    wb.save(xlsx_path)
    if merged_count or conflict_count:
        print(f'Merged Import -> Metadata: {merged_count} cell(s) filled, {conflict_count} conflict(s) resolved.')


def load_metadata_tool_module():
    spec_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'lrplugin-dev', 'metadataTool.py')
    spec = importlib.util.spec_from_file_location('lr_metadata_tool', spec_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def idx_label(index):
    return str(index).zfill(3) if index is not None else '???'


def roll_root_from_raw_path(raw_path):
    # A selected photo's raw file path looks like eg.
    # "{LIBRARY_PATH}/012_2024-01-01_STK_CAM_Name/01_scans/DSC001.ARW" --
    # the roll folder is just the first path segment below LIBRARY_PATH.
    # Returns None if raw_path isn't under LIBRARY_PATH at all (a photo from
    # some other part of the catalog).
    try:
        rel = os.path.relpath(raw_path, LIBRARY_PATH)
    except ValueError:
        return None
    if rel.startswith('..') or rel == '.':
        return None
    top = rel.split(os.sep)[0]
    candidate = os.path.join(LIBRARY_PATH, top)
    return candidate if os.path.isdir(candidate) else None


def detect_roll_from_lr_selection(lr_module):
    """Best-effort auto-detect: reads whatever's CURRENTLY selected in
    Lightroom (no selection change is made) via lr_module.read_lr_selection(),
    and if every selected photo's raw file resolves to the SAME roll folder
    under LIBRARY_PATH, returns (roll_root, index) for it -- skipping the
    manual index prompt entirely. Falls back to (None, None) -- meaning
    prompt_roll_folder() should run instead -- on anything ambiguous:
    nothing selected, plugin not enabled, photos outside LIBRARY_PATH, or a
    selection spanning more than one roll."""
    rows = lr_module.read_lr_selection()
    if not rows:
        return None, None

    roll_roots = {roll_root_from_raw_path(path) for _fname, path in rows}
    roll_roots.discard(None)

    if len(roll_roots) == 0:
        print("Current Lightroom selection isn't under LIBRARY_PATH -- enter the roll index manually.")
        return None, None
    if len(roll_roots) > 1:
        print(f'Current Lightroom selection spans {len(roll_roots)} different rolls -- enter the roll index manually.')
        return None, None

    roll_root = roll_roots.pop()
    index = roll_index_from_folder_name(roll_root)
    print(f'Auto-detected roll from current Lightroom selection: {os.path.basename(roll_root)}')
    return roll_root, index


def main():
    print(f'Importing metadata for library: {LIBRARY_PATH}')

    lr_module = load_metadata_tool_module()

    roll_root, index = detect_roll_from_lr_selection(lr_module)
    if roll_root is None:
        roll_root, index = prompt_roll_folder()

    activate_lightroom()

    scans_path = os.path.join(roll_root, '01_scans')

    xlsx_path = find_metadata_xlsx(roll_root)
    if xlsx_path is None:
        print(f'No *_metadata.xlsx found in {roll_root}. Run newRoll.py for this roll first.')
        return

    collection = collectionObj.collectionObj(LIBRARY_PATH)

    ensure_workbook_structure(xlsx_path)
    merge_import_into_metadata(xlsx_path, collection)

    missing_on_disk, missing_in_xlsx, xlsx_count, disk_count = reconcile(xlsx_path, scans_path)

    if missing_on_disk or missing_in_xlsx:
        print(f'\n[{idx_label(index)}] Metadata xlsx and 01_scans are out of sync:')
        print(f'  xlsx rows: {xlsx_count}')
        print(f'  raw files: {disk_count}')
        if missing_on_disk:
            print(f'  in xlsx but no longer on disk ({len(missing_on_disk)}): {missing_on_disk}')
        if missing_in_xlsx:
            print(f'  on disk but missing from xlsx ({len(missing_in_xlsx)}): {missing_in_xlsx}')

        print("\nApplying metadata positionally with a mismatch risks pasting the wrong")
        print("frame's data onto a photo in Lightroom. Refusing to proceed by default.")

        choice = input('Auto-fix the xlsx (drop stale rows / add missing rows) and re-check? [y/N] ').strip().lower()
        if choice not in ('y', 'yes'):
            print('Aborted. Lightroom was not touched. Fix the xlsx by hand and rerun.')
            return

        auto_fix_xlsx(xlsx_path, scans_path, roll_root, collection)

        missing_on_disk, missing_in_xlsx, xlsx_count, disk_count = reconcile(xlsx_path, scans_path)
        if missing_on_disk or missing_in_xlsx:
            print('Still mismatched after auto-fix -- aborting. Lightroom was not touched.')
            return

    print(f'\n[{idx_label(index)}] xlsx and 01_scans match ({xlsx_count} files). Proceeding to Lightroom import.')

    MetadataTool = lr_module.metadataTool
    tool = MetadataTool(xlsx_path=xlsx_path, raw_folder=scans_path)
    tool.pause_field = False
    tool.pause_nextImage = False
    tool.run()


if __name__ == '__main__':
    main()
