# newRoll.py
#
# Generates a template folder for a newly-scanned roll, before any JPGs/EXIF exist yet --
# the gap in the current pipeline where a freshly scanned roll has nowhere to go.
#
# Workflow:
#   1) prompts for STK (film stock) and CAM (camera), strictly validated against
#      data/stocklist.xlsx / data/cameralist.xlsx (same lookup tables collectionObj uses),
#      and a free-text Name/location tag
#   2) builds {index}_{YY-MMs-MMe}_{STK}_{CAM}_{Name} folder structure matching the
#      "new collection" layout rollObj.find_image_dirs() already expects
#      (01_scans / 02_exports / 04_edits / 05_other/01_unmatched_raws), using today's
#      date for both start+end month (matches the {YY-MMs-MMe} convention cleanRoll's
#      later rename uses, just collapsed to a single day since no exposures exist yet)
#   3) opens 01_scans in Finder and waits for you to copy raw files in
#   4) builds a per-frame metadata.xlsx template (same schema as
#      lrplugin-dev/metadataTool.py's generate_template()) for the raw files found,
#      prefilled with the roll-level STK/CAM/lab-name you already entered plus a set of
#      hardcoded roll-wide defaults (scan equipment, light source, film holder by format)
#
# Usage:
#   python newRoll.py

import os
import re
import subprocess
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

import collectionObj

# Working "scan/edit" library root -- where rolls live while you're still in Lightroom
# editing them, before cleanRoll.py archives them into the final dated library.
# Matches main.py's non-DEVMODE `library` constant. Edit as needed.
LIBRARY_PATH = r'/Users/rja/Photography/0_Working/1_Imports/'

RAW_EXTS = ('.arw', '.dng', '.tif', '.tiff')

# Negative Lab Pro's "Create Positive .tiff" + "Stack with Original" option
# writes a "{base}-positive.tif" file into the same 01_scans folder as the
# raw capture it was generated from, stacked with it in Lightroom. It's a
# stack companion of an already-counted exposure, not its own -- list_raw_files()
# skips anything matching this pattern so it never gets counted as, or given,
# its own xlsx row.
# Negative Lab Pro's "Create Positive .tiff" + "Stack with Original" option
# writes a "{base}-positive.tif" file into the same 01_scans folder as the
# raw capture it was generated from, stacked with it in Lightroom. It's a
# stack companion of an already-counted exposure, not its own -- list_raw_files()
# skips anything matching this pattern so it never gets counted as, or given,
# its own xlsx row.
#
# Matches BOTH "-positive" and "_positive": "-" sorts before "." in any
# filename-sorted view (eg. Quick Collection sorted by filename), which puts
# "{base}-positive.tif" BEFORE its own master's "{base}.ARW" -- breaking
# syncVCs.py's assumption that the master is always visually first in its
# stack. Renaming to "_positive.tif" (underscore sorts after ".") fixes
# that; matching both here means this keeps working whether a given file
# has been renamed yet or not.
#
# Also matches an optional "-N" suffix ("-positive-2.tif", "-positive-3.tif",
# ...) -- NLP can generate more than one positive per raw capture, and each
# additional one gets a numbered suffix (confirmed against real data:
# "DSC00002-positive.tif" + "DSC00002-positive-2.tif" both existing for the
# same raw).
NLP_POSITIVE_TIFF_RE = re.compile(r'[-_]positive(-\d+)?\.tiff?$', re.IGNORECASE)

# Roll-wide metadata defaults, hardcoded per your fixed workflow. Edit these if your
# gear/lab setup changes.
SCAN_EQUIPMENT = 'AS-2'
LIGHT_SOURCE = 'Cinelite'
FILM_HOLDER_135 = 'AS-135-2'
FILM_HOLDER_120 = 'AS-120-2'

METADATA_COLUMNS = [
    "Index", "rawFileName", "rawFilePath",
    "Year", "Month", "Day",
    "Sublocation", "City", "State", "Country/Region",
    "Intellectual Genre", "Scene",
    "Camera Make", "Camera Model", "Lens Make", "Lens Model",
    "Film Stock", "Film ISO", "Gear Notes",
    "Shot at ISO", "Aperture", "Shutter Speed", "Focal Length", "Shooting Notes",
    "Scan Equipment", "Light Source", "Film Holder", "Digitization Notes",
    "Developer", "Dilution", "Dev Time/Temp", "Dev Method", "Dev Notes",
]

# Condensed, user-facing subset of METADATA_COLUMNS for the "Import" sheet --
# the fields a user actually fills in by hand, grouped and colored so the
# sheet is easy to work in. importMetadata.py merges these into the
# "Metadata" sheet (row-matched by rawFileName) on the next run.
#
# "Notes" is Import-only and never merges into Metadata -- it's a personal
# scratch field for filling out the sheet, not part of the schema
# lrplugin-dev/metadataTool.py hands to Lightroom.
IMPORT_COLUMNS = [
    "Notes", "Index", "rawFileName",
    "Year", "Month", "Day",
    "Sublocation", "City", "State", "Country/Region",
    "Lens Make", "Lens Model",
    "Aperture", "Shutter Speed", "Focal Length",
]

IMPORT_COLUMN_GROUPS = [
    (("Notes", "Index", "rawFileName"), "DCE6F1"),                 # a
    (("Year", "Month", "Day"), "E2EFDA"),                           # b
    (("Sublocation", "City", "State", "Country/Region"), "FCE4D6"), # c
    (("Lens Make", "Lens Model"), "E4DFEC"),                        # d
    (("Aperture", "Shutter Speed", "Focal Length"), "FFF2CC"),      # e
]


def force_text_format(ws, column_names=None):
    """Sets column-level number_format='@' (Text) so Excel never silently
    reinterprets a typed value like "1/15" or "250" as a date/number --
    applies to any cell in that column, including rows added later, since
    it's stored as the column's own default format rather than per-cell."""
    header = [cell.value for cell in ws[1]]
    for idx, name in enumerate(header, start=1):
        if column_names is not None and name not in column_names:
            continue
        ws.column_dimensions[get_column_letter(idx)].number_format = '@'


def style_import_sheet(ws):
    header = [cell.value for cell in ws[1]]
    for names, hex_color in IMPORT_COLUMN_GROUPS:
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        for idx, name in enumerate(header, start=1):
            if name in names:
                ws.column_dimensions[get_column_letter(idx)].fill = fill
                ws.cell(row=1, column=idx).font = Font(bold=True)
    force_text_format(ws)


def build_import_sheet(wb, rows):
    """rows: list of (index, raw_file_name). Creates (or replaces) the
    Import sheet, prefilled with Index/rawFileName so it's row-aligned with
    Metadata from the start."""
    if 'Import' in wb.sheetnames:
        del wb['Import']
    ws = wb.create_sheet('Import')
    ws.append(IMPORT_COLUMNS)
    idx_col = IMPORT_COLUMNS.index('Index')
    name_col = IMPORT_COLUMNS.index('rawFileName')
    for index, raw_file_name in rows:
        row = [None] * len(IMPORT_COLUMNS)
        row[idx_col] = index
        row[name_col] = raw_file_name
        ws.append(row)
    style_import_sheet(ws)
    return ws


def get_next_index(library_path):
    max_idx = 0
    if os.path.isdir(library_path):
        for name in os.listdir(library_path):
            if name.startswith('.'):
                continue
            full = os.path.join(library_path, name)
            if not os.path.isdir(full):
                continue
            token = name.split('_')[0] if '_' in name else name.split(' - ')[0]
            token = token.strip()
            if token.isdigit():
                max_idx = max(max_idx, int(token))
    return max_idx + 1


def prompt_index(default_idx):
    raw = input(f'Roll index [{default_idx}]: ').strip()
    if raw == '':
        return default_idx
    if not raw.isdigit():
        print('Index must be a number, using default.')
        return default_idx
    return int(raw)


def build_stock_index(stocklist):
    index = {}
    for entry in stocklist.values():
        for key in (entry.get('KEY_ID'), entry.get('stk'), entry.get('stock')):
            if key:
                index[key.strip().lower()] = entry
    return index


def prompt_stock(collection):
    index = build_stock_index(collection.stocklist)
    while True:
        raw = input('Film stock (STK code / stock name): ').strip()
        if not raw:
            print('Stock is required.')
            continue
        entry = index.get(raw.lower())
        if entry:
            return entry
        matches = [v for k, v in index.items() if raw.lower() in k]
        if matches:
            names = sorted({m['stock'] for m in matches if m.get('stock')})
            print(f'No exact match in stocklist.xlsx. Did you mean: {", ".join(names)}')
        else:
            print(f'"{raw}" not found in stocklist.xlsx. Add it there first, or check spelling.')


def load_camera_rows():
    """Reads cameralist.xlsx rows directly, instead of going through
    collectionObj.build_cameralist()'s dict. That dict is keyed by id/model/"brand model"
    with a plain assignment per row, so when multiple distinct camera bodies share the
    same model name (eg. three "F3"s with ids F3/F3S/F3'), each later row silently
    overwrites the earlier one under the shared 'model' key -- whichever row happens to
    be listed last in the sheet wins, with no way to tell from the result. Reading the
    raw rows here lets prompt_camera() detect that ambiguity and ask, instead of
    guessing."""
    project_dir = os.path.dirname(os.path.abspath(collectionObj.__file__))
    xlsx_path = os.path.join(project_dir, 'data', 'cameralist.xlsx')
    df = pd.read_excel(xlsx_path, dtype=str, engine='openpyxl').fillna('')

    rows = []
    for _, row in df.iterrows():
        cam_id = row['id'].strip()
        if not cam_id:
            continue
        rows.append({
            'id': cam_id,
            'model': row['model'].strip(),
            'brand': row['brand'].strip(),
            'serial': row['serial'].strip(),
            'filmtype': row['filmtype'].strip(),
            'filmformat': row['filmformat'].strip(),
        })
    return rows


def prompt_camera():
    rows = load_camera_rows()

    while True:
        raw = input('Camera (ID / model / "brand model"): ').strip()
        if not raw:
            print('Camera is required.')
            continue

        key = raw.lower()
        matches = [
            r for r in rows
            if key in (r['id'].lower(), r['model'].lower(), f"{r['brand']} {r['model']}".strip().lower())
        ]

        if not matches:
            loose = [r for r in rows if key in r['model'].lower() or key in r['id'].lower()]
            if loose:
                names = sorted({f"{r['brand']} {r['model']} (id={r['id']})" for r in loose})
                print(f'No exact match in cameralist.xlsx. Did you mean: {", ".join(names)}')
            else:
                print(f'"{raw}" not found in cameralist.xlsx. Add it there first, or check spelling.')
            continue

        # Collapse true duplicate rows (identical id/model/brand/filmtype -- harmless
        # repeated lines in the sheet) before checking for genuine ambiguity.
        unique = {}
        for r in matches:
            unique[(r['id'], r['model'], r['brand'], r['filmtype'])] = r
        matches = list(unique.values())

        if len(matches) == 1:
            return matches[0]

        # Genuinely ambiguous: the typed name matches more than one distinct camera body
        # (eg. several "F3"s with different ids/serials). Force a pick instead of
        # silently taking whichever row cameralist.xlsx happens to list last -- this is
        # exactly the bug that produced an unexpected "F3S" for a typed "F3".
        print(f'"{raw}" matches {len(matches)} cameras in cameralist.xlsx -- pick one:')
        for i, r in enumerate(matches, start=1):
            serial = f", serial={r['serial']}" if r['serial'] else ''
            print(f"  {i}) id={r['id']:<10} {r['brand']} {r['model']}  ({r['filmtype']}/{r['filmformat']}{serial})")

        while True:
            pick = input('Number: ').strip()
            if pick.isdigit() and 1 <= int(pick) <= len(matches):
                return matches[int(pick) - 1]
            print('Not a valid choice, try again.')


def prompt_name():
    while True:
        raw = input('Name / location tag for this roll (eg. "Zurich Flims"): ').strip()
        if raw:
            return raw
        print('Name is required.')


def prompt_lab_name():
    while True:
        raw = input('Lab name: ').strip()
        if raw:
            return raw
        print('Lab name is required.')


def resolve_film_holder(cam_entry):
    # cameralist.xlsx's `filmtype` is the base spool format (135/120/45/810), distinct
    # from `filmformat` (the frame geometry, eg. 6x6/6x7 -- see renderTool.FORMATS).
    filmtype = (cam_entry.get('filmtype') or '').strip()
    if filmtype == '135':
        return FILM_HOLDER_135
    if filmtype == '120':
        return FILM_HOLDER_120
    return None


def build_folder_name(index, stk_entry, cam_entry, name):
    today = datetime.now()
    # {YY-MMs-MMe}: start month == end month == today's month, since no exposures exist yet.
    # cleanRoll.py's own rename pass overwrites this with the real shoot dates later.
    date_str = f"{today.strftime('%y')}-{today.strftime('%m')}-{today.strftime('%m')}"
    index_str = str(index).zfill(3)
    stk_code = stk_entry['stk']
    cam_id = cam_entry['id']
    safe_name = name.replace('/', '-').replace('_', '-')
    return f"{index_str}_{date_str}_{stk_code}_{cam_id}_{safe_name}"


def build_roll_structure(library_path, folder_name):
    root = os.path.join(library_path, folder_name)
    dirs = [
        root,
        os.path.join(root, '01_scans'),
        os.path.join(root, '02_exports'),
        os.path.join(root, '04_edits'),
        os.path.join(root, '05_other', '01_unmatched_raws'),
    ]
    for d in dirs:
        os.makedirs(d, exist_ok=True)
    return root


def wait_for_raw_files(scans_path):
    try:
        subprocess.run(['open', scans_path])
    except Exception:
        pass
    input(f'\nCopy your RAW/scan files into:\n  {scans_path}\nPress Enter once done (or to skip and add them later)...\n')


def list_raw_files(scans_path):
    files = []
    for name in sorted(os.listdir(scans_path)):
        if name.startswith('.'):
            continue
        if NLP_POSITIVE_TIFF_RE.search(name):
            continue
        if name.lower().endswith(RAW_EXTS):
            files.append(name)
    return files


def build_metadata_template(roll_root, folder_name, raw_files, stk_entry, cam_entry, lab_name):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Metadata'
    ws.append(METADATA_COLUMNS)

    film_holder = resolve_film_holder(cam_entry)

    for i, fname in enumerate(raw_files, start=1):
        row = [None] * len(METADATA_COLUMNS)
        row[0] = i                              # Index
        row[1] = fname                          # rawFileName
        row[2] = os.path.join(roll_root, '01_scans', fname)  # rawFilePath

        # Corrected mapping (was swapped relative to your established
        # existing rolls -- verified against real historical files): Scene
        # holds the stock code, Intellectual Genre holds camera(+lens).
        row[10] = cam_entry.get('id')           # Intellectual Genre -- CAM for now, CAM+LNS later
        # Intellectual Genre should eventually be CAM + LNS (eg. "F3 28mm2.8"),
        # but no lens ID exists yet at roll-creation time (lens is only known
        # once EXIF exists) -- prefill with just CAM for now, matching your
        # existing filled-out rolls, and append LNS in a later pass once
        # frames have lens data (see importMetadata.py / metadataTool.py).
        row[11] = stk_entry.get('stk')           # Scene -- matches STK

        row[12] = cam_entry.get('brand')        # Camera Make
        row[13] = cam_entry.get('model')        # Camera Model

        row[16] = stk_entry.get('stock')        # Film Stock
        row[17] = stk_entry.get('boxspeed')     # Film ISO
        row[19] = stk_entry.get('boxspeed')     # Shot at ISO -- matches ISO (box speed by default; edit per-row if pushed/pulled)

        row[24] = SCAN_EQUIPMENT                # Scan Equipment
        row[25] = LIGHT_SOURCE                  # Light Source
        row[26] = film_holder                   # Film Holder (135/120, by camera's filmtype)

        row[28] = stk_entry.get('process')      # Developer -- matches stock's process (C41/E6/BNW)
        row[32] = lab_name                      # Dev Notes -- lab name

        ws.append(row)

    force_text_format(ws, IMPORT_COLUMNS)   # text-format the columns shared with Import here too
    build_import_sheet(wb, [(i, fname) for i, fname in enumerate(raw_files, start=1)])
    wb.active = wb.sheetnames.index('Metadata')   # keep Metadata active regardless of sheet-creation order

    out_path = os.path.join(roll_root, f'{folder_name}_metadata.xlsx')
    wb.save(out_path)
    return out_path


def main():
    collection = collectionObj.collectionObj(LIBRARY_PATH)

    next_idx = get_next_index(LIBRARY_PATH)
    index = prompt_index(next_idx)

    stk_entry = prompt_stock(collection)
    cam_entry = prompt_camera()
    name = prompt_name()
    lab_name = prompt_lab_name()

    if resolve_film_holder(cam_entry) is None:
        print(f'Note: camera filmtype "{cam_entry.get("filmtype")}" has no Film Holder default.')

    folder_name = build_folder_name(index, stk_entry, cam_entry, name)
    roll_root = build_roll_structure(LIBRARY_PATH, folder_name)
    scans_path = os.path.join(roll_root, '01_scans')

    print(f'\nCreated roll template:\n  {roll_root}\n')

    wait_for_raw_files(scans_path)

    raw_files = list_raw_files(scans_path)
    if not raw_files:
        print('No RAW files found in 01_scans yet -- metadata template will have no rows. '
              'Re-run newRoll.py\'s build_metadata_template() later once scans are copied in.')

    metadata_path = build_metadata_template(roll_root, folder_name, raw_files, stk_entry, cam_entry, lab_name)

    print(f'\nRoll {str(index).zfill(3)} ready:')
    print(f'  folder:   {roll_root}')
    print(f'  scans:    {scans_path}  ({len(raw_files)} raw files found)')
    print(f'  metadata: {metadata_path}')
    print('\nNext: import into Lightroom, edit, fill in the metadata xlsx, run '
          'lrplugin-dev/metadataTool.py, export JPGs into 02_exports, then run cleanRoll.py.')


if __name__ == '__main__':
    main()
